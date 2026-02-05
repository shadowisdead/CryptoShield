"""
File metadata management with optional encrypted storage.
Supports extended records, search, filter, and export.
"""

import csv
import json
import os
from dataclasses import dataclass, asdict
from typing import Optional

try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM #type: ignore
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC #type: ignore
    from cryptography.hazmat.primitives import hashes #type: ignore
    from cryptography.hazmat.backends import default_backend #type: ignore
    HAS_CRYPTO = True
except ImportError:
    HAS_CRYPTO = False

DATA_DIR = "data"
DATA_FILE = os.path.join(DATA_DIR, "metadata.json")
DATA_FILE_ENC = os.path.join(DATA_DIR, "metadata.enc")


@dataclass
class FileRecord:
    original_file: str
    encrypted_file: str
    file_hash: str
    time: str
    algorithm: str = "AES-256"
    file_size: int = 0
    signature: Optional[str] = None

    def to_dict(self) -> dict:
        d = {
            "original_file": self.original_file,
            "encrypted_file": self.encrypted_file,
            "file_hash": self.file_hash,
            "time": self.time,
            "algorithm": self.algorithm,
            "file_size": self.file_size,
        }
        if self.signature:
            d["signature"] = self.signature
        return d

    @classmethod
    def from_dict(cls, d: dict) -> "FileRecord":
        return cls(
            original_file=d.get("original_file", ""),
            encrypted_file=d.get("encrypted_file", ""),
            file_hash=d.get("file_hash", ""),
            time=d.get("time", ""),
            algorithm=d.get("algorithm", "AES-256"),
            file_size=int(d.get("file_size", 0)),
            signature=d.get("signature"),
        )


class FileManager:
    """Manages encrypted file metadata with optional encrypted storage."""

    def __init__(self, metadata_password: Optional[str] = None):
        os.makedirs(DATA_DIR, exist_ok=True)
        self._metadata_password = metadata_password
        self._salt_path = os.path.join(DATA_DIR, ".salt")

    def _get_key(self) -> Optional[bytes]:
        """Derive encryption key for metadata from password."""
        if not self._metadata_password or not HAS_CRYPTO:
            return None
        if os.path.exists(self._salt_path):
            with open(self._salt_path, "rb") as f:
                salt = f.read()
        else:
            salt = os.urandom(16)
            with open(self._salt_path, "wb") as f:
                f.write(salt)
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=600_000,
            backend=default_backend(),
        )
        return kdf.derive(self._metadata_password.encode("utf-8"))

    def _encrypt_json(self, data: str) -> bytes:
        key = self._get_key()
        if not key:
            return data.encode("utf-8")
        aesgcm = AESGCM(key)
        nonce = os.urandom(12)
        enc = aesgcm.encrypt(nonce, data.encode("utf-8"), None)
        return nonce + enc

    def _decrypt_json(self, raw: bytes) -> str:
        key = self._get_key()
        if not key or len(raw) < 13:
            return raw.decode("utf-8", errors="replace")
        aesgcm = AESGCM(key)
        nonce, ct = raw[:12], raw[12:]
        return aesgcm.decrypt(nonce, ct, None).decode("utf-8")

    def _read_records(self) -> list:
        if self._get_key() and os.path.exists(DATA_FILE_ENC):
            try:
                with open(DATA_FILE_ENC, "rb") as f:
                    raw = f.read()
                text = self._decrypt_json(raw)
                return json.loads(text)
            except Exception:
                pass
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return []

    def _write_records(self, records: list) -> None:
        text = json.dumps(records, indent=2, ensure_ascii=False)
        if self._get_key():
            enc = self._encrypt_json(text)
            with open(DATA_FILE_ENC, "wb") as f:
                f.write(enc)
            if os.path.exists(DATA_FILE):
                os.remove(DATA_FILE)
        else:
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                f.write(text)
            if os.path.exists(DATA_FILE_ENC):
                os.remove(DATA_FILE_ENC)

    def save_record(self, record: FileRecord) -> None:
        records = self._read_records()
        records.append(record.to_dict() if hasattr(record, "to_dict") else asdict(record))
        self._write_records(records)

    def get_all_records(self) -> list:
        """Returns list of dicts for compatibility with GUI."""
        raw = self._read_records()
        result = []
        for r in raw:
            if isinstance(r, dict):
                d = dict(r)
                d.setdefault("algorithm", "AES-256")
                d.setdefault("file_size", 0)
                result.append(d)
            else:
                result.append(FileRecord.from_dict(r).to_dict())
        return result

    def search_records(
        self,
        filename: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        hash_substring: Optional[str] = None,
        algorithm: Optional[str] = None,
    ) -> list:
        """Search and filter records by filename, date range, hash, or algorithm."""
        records = self.get_all_records()
        result = []
        for r in records:
            rd = r if isinstance(r, dict) else (r.to_dict() if hasattr(r, "to_dict") else asdict(r))
            if filename and filename.lower() not in rd.get("original_file", "").lower():
                continue
            if hash_substring and hash_substring.lower() not in rd.get("file_hash", "").lower():
                continue
            if algorithm and rd.get("algorithm", "").lower() != algorithm.lower():
                continue
            t = rd.get("time", "")
            if date_from and t < date_from:
                continue
            if date_to and t > date_to:
                continue
            result.append(rd)
        return result

    def export_csv(self, path: str, records: Optional[list] = None) -> None:
        """Export records to CSV file."""
        recs = records or [r.to_dict() if hasattr(r, "to_dict") else asdict(r) for r in self.get_all_records()]
        if not recs:
            return
        cols = list(recs[0].keys())
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(recs)

    def export_excel(self, path: str, records: Optional[list] = None) -> bool:
        """Export records to Excel (.xlsx). Returns True if successful."""
        try:
            import openpyxl #type: ignore
        except ImportError:
            return False
        recs = records or [r.to_dict() if hasattr(r, "to_dict") else asdict(r) for r in self.get_all_records()]
        if not recs:
            return True
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "History"
        cols = list(recs[0].keys())
        for c, col in enumerate(cols, 1):
            ws.cell(row=1, column=c, value=col)
        for r, row in enumerate(recs, 2):
            for c, col in enumerate(cols, 1):
                ws.cell(row=r, column=c, value=row.get(col, ""))
        wb.save(path)
        return True
